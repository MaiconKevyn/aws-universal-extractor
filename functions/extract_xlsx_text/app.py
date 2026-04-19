import io
from typing import Any

from openpyxl import load_workbook

from app_common.exceptions import DocumentExtractionError
from app_common.logging import get_logger, log_json
from app_common.s3_utils import get_object_bytes, put_text


logger = get_logger(__name__)


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.2f}"
    return str(value)


def _sheet_to_text(sheet) -> str:
    lines: list[str] = []
    for row in sheet.iter_rows(values_only=True):
        cells = [_cell_to_str(c) for c in row]
        if not any(c.strip() for c in cells):
            continue
        lines.append("| " + " | ".join(cells) + " |")
    return "\n".join(lines)


def lambda_handler(event: dict[str, Any], context: Any) -> dict[str, Any]:
    bucket = event["document"]["bucket"]
    key = event["document"]["key"]
    output_bucket = event["artifacts"]["output_bucket"]
    output_prefix = event["artifacts"]["output_prefix"]

    xlsx_bytes = get_object_bytes(bucket, key)
    workbook = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)

    sections: list[str] = []
    sheet_summary: list[dict[str, Any]] = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_text = _sheet_to_text(sheet)
        if not sheet_text.strip():
            continue
        sections.append(f"=== Sheet: {sheet_name} ===\n{sheet_text}")
        sheet_summary.append({
            "name": sheet_name,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
        })

    document_text = "\n\n".join(sections)

    if not document_text.strip():
        raise DocumentExtractionError("XLSX workbook is empty - no cells contain data")

    raw_text_key = f"{output_prefix}/raw_text.txt"
    put_text(output_bucket, raw_text_key, document_text)

    event["artifacts"]["raw_text"] = {
        "bucket": output_bucket,
        "key": raw_text_key,
    }
    event["xlsx_extraction"] = {
        "engine": "openpyxl",
        "sheets": sheet_summary,
        "text_length": len(document_text),
    }

    log_json(
        logger,
        "XLSX text extracted",
        request_id=event["request_id"],
        sheet_count=len(sheet_summary),
        text_length=len(document_text),
    )
    return event
