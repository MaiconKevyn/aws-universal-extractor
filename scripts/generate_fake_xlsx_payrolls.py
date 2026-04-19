"""Generate synthetic US payroll (pay stub) XLSX spreadsheets for testing the pipeline.

Dev-only dependencies (not shipped to Lambda):
    pip install openpyxl faker

Usage:
    python scripts/generate_fake_xlsx_payrolls.py --out tests/fixtures/payroll/xlsx --count 5

Each XLSX is paired with a <stem>.expected.json ground-truth file — the SAME schema
used by the PDF fixtures, since the `payroll/v1` profile is format-agnostic.
"""

from __future__ import annotations

import argparse
import json
import random
import sys
from pathlib import Path

from faker import Faker
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _paystub_data import Paystub, build_paystub, paystub_to_ground_truth  # noqa: E402


HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
SECTION_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
HIGHLIGHT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
BOLD = Font(bold=True)
TITLE = Font(bold=True, size=14)


def render_xlsx(p: Paystub, out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Paystub"

    row = 1

    ws.cell(row=row, column=1, value=p.employer_name).font = TITLE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    ws.cell(row=row, column=1, value=p.employer_address)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    ws.cell(row=row, column=1, value=f"EIN: {p.employer_ein}")
    row += 2

    ws.cell(row=row, column=1, value="EARNINGS STATEMENT").font = BOLD
    ws.cell(row=row, column=4, value="Pay Date:").font = BOLD
    ws.cell(row=row, column=5, value=p.pay_date)
    row += 2

    info_rows = [
        ("Employee", p.employee_name, "Employee ID", p.employee_id),
        ("Job Title", p.job_title, "SSN", f"XXX-XX-{p.ssn_last4}"),
        ("Hire Date", p.hire_date, "Pay Frequency", p.pay_frequency.title()),
        ("Pay Period", f"{p.period_start} to {p.period_end}", "Pay Rate", p.pay_rate),
    ]
    for label1, val1, label2, val2 in info_rows:
        ws.cell(row=row, column=1, value=label1).font = BOLD
        ws.cell(row=row, column=2, value=val1)
        ws.cell(row=row, column=4, value=label2).font = BOLD
        ws.cell(row=row, column=5, value=val2)
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.cell(row=row, column=4).fill = SECTION_FILL
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Earnings").font = BOLD
    row += 1
    earn_headers = ["Code", "Description", "Hours", "Rate", "Current", "YTD"]
    for col, h in enumerate(earn_headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = BOLD
        cell.fill = HEADER_FILL
    row += 1
    for item in (i for i in p.line_items if i.kind == "earning"):
        ws.cell(row=row, column=1, value=item.code)
        ws.cell(row=row, column=2, value=item.description)
        ws.cell(row=row, column=3, value=item.hours)
        ws.cell(row=row, column=4, value=item.rate)
        c5 = ws.cell(row=row, column=5, value=item.amount)
        c6 = ws.cell(row=row, column=6, value=item.ytd_amount)
        c5.number_format = '#,##0.00'
        c6.number_format = '#,##0.00'
        row += 1
    ws.cell(row=row, column=2, value="Gross Pay").font = BOLD
    c5 = ws.cell(row=row, column=5, value=p.gross_pay)
    c6 = ws.cell(row=row, column=6, value=p.ytd_gross_pay)
    c5.number_format = c6.number_format = '#,##0.00'
    c5.font = c6.font = BOLD
    for col in range(1, 7):
        ws.cell(row=row, column=col).fill = SECTION_FILL
    row += 2

    ws.cell(row=row, column=1, value="Deductions").font = BOLD
    row += 1
    ded_headers = ["Code", "Description", "", "", "Current", "YTD"]
    for col, h in enumerate(ded_headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = BOLD
        if h:
            cell.fill = HEADER_FILL
    row += 1
    for item in (i for i in p.line_items if i.kind == "deduction"):
        ws.cell(row=row, column=1, value=item.code)
        ws.cell(row=row, column=2, value=item.description)
        c5 = ws.cell(row=row, column=5, value=item.amount)
        c6 = ws.cell(row=row, column=6, value=item.ytd_amount)
        c5.number_format = '#,##0.00'
        c6.number_format = '#,##0.00'
        row += 1
    ws.cell(row=row, column=2, value="Total Deductions").font = BOLD
    c5 = ws.cell(row=row, column=5, value=p.total_deductions)
    c6 = ws.cell(row=row, column=6, value=p.ytd_gross_pay - p.ytd_net_pay)
    c5.number_format = c6.number_format = '#,##0.00'
    c5.font = c6.font = BOLD
    for col in range(1, 7):
        ws.cell(row=row, column=col).fill = SECTION_FILL
    row += 2

    summary = [
        ("Gross Pay", p.gross_pay, "YTD Gross", p.ytd_gross_pay),
        ("Total Taxes", p.total_taxes, "Total Deductions", p.total_deductions),
        ("", "", "NET PAY", p.net_pay),
    ]
    for label1, val1, label2, val2 in summary:
        ws.cell(row=row, column=1, value=label1).font = BOLD
        if isinstance(val1, (int, float)):
            c = ws.cell(row=row, column=2, value=val1)
            c.number_format = '#,##0.00'
        ws.cell(row=row, column=4, value=label2).font = BOLD
        if isinstance(val2, (int, float)):
            c = ws.cell(row=row, column=5, value=val2)
            c.number_format = '#,##0.00'
            if label2 == "NET PAY":
                c.fill = HIGHLIGHT_FILL
                c.font = BOLD
                ws.cell(row=row, column=4).fill = HIGHLIGHT_FILL
        row += 1

    widths = {1: 18, 2: 32, 3: 10, 4: 18, 5: 14, 6: 14}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left")

    wb.save(out_path)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", type=Path, required=True)
    ap.add_argument("--count", type=int, default=5)
    ap.add_argument("--seed", type=int, default=42)
    ap.add_argument("--locale", default="en_US")
    ap.add_argument(
        "--variant",
        choices=("canonical", "with_overtime", "with_bonus", "mixed"),
        default="mixed",
    )
    args = ap.parse_args()

    random.seed(args.seed)
    fake = Faker(args.locale)
    Faker.seed(args.seed)

    args.out.mkdir(parents=True, exist_ok=True)

    variants = (
        ["canonical", "with_overtime", "with_bonus"]
        if args.variant == "mixed"
        else [args.variant]
    )

    for i in range(args.count):
        variant = variants[i % len(variants)]
        p = build_paystub(fake, variant=variant)
        stem = f"paystub_{i + 1:03d}_{variant}"
        xlsx_path = args.out / f"{stem}.xlsx"
        gt_path = args.out / f"{stem}.expected.json"

        render_xlsx(p, xlsx_path)
        gt_path.write_text(
            json.dumps(paystub_to_ground_truth(p), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(f"wrote {xlsx_path}")


if __name__ == "__main__":
    main()
