"""Production evaluation harness for local fixtures.

Offline mode is deterministic and CI-safe: validates fixture ground truth,
normalizes each document to text, and checks schema contracts.

LLM mode additionally calls OpenAI, compares field-level accuracy against
<stem>.expected.json, and enforces a threshold. CI can run this when
OPENAI_API_KEY is available as a repository secret.
"""

from __future__ import annotations

import argparse
import contextlib
import io
import json
import math
import sys
from pathlib import Path
from typing import Any, Callable

from docx import Document
from jsonschema import Draft202012Validator
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT / "layers" / "common" / "python"))
sys.path.insert(0, str(REPO_ROOT))

from app_common.openai_client import OpenAIExtractionClient  # noqa: E402
from app_common.profiles import load_profile  # noqa: E402
from app_common.prompt_safety import wrap_untrusted_document_text  # noqa: E402
from app_common.validators import to_metadata_json  # noqa: E402
from functions.extract_csv_text.app import _decode_csv, _rows_to_text, _sniff_dialect  # noqa: E402
from functions.extract_docx_text.app import _paragraph_text, _table_to_text  # noqa: E402
from functions.extract_pdf_text.classifier import classify  # noqa: E402
from functions.extract_pdf_text.strategies import extract_text_layer  # noqa: E402
from functions.extract_xlsx_text.app import _sheet_to_text  # noqa: E402


Extractor = Callable[[Path], str]


def extract_pdf(path: Path) -> str:
    with contextlib.redirect_stdout(io.StringIO()), contextlib.redirect_stderr(io.StringIO()):
        pages = extract_text_layer(path.read_bytes())
    return "\n\n".join(f"=== Page {p.page_number} ({p.method}) ===\n{p.markdown}" for p in pages)


def extract_xlsx(path: Path) -> str:
    workbook = load_workbook(io.BytesIO(path.read_bytes()), read_only=True, data_only=True)
    sections: list[str] = []
    for sheet_name in workbook.sheetnames:
        text = _sheet_to_text(workbook[sheet_name])
        if text.strip():
            sections.append(f"=== Sheet: {sheet_name} ===\n{text}")
    return "\n\n".join(sections)


def extract_csv(path: Path) -> str:
    text = _decode_csv(path.read_bytes())
    dialect = _sniff_dialect(text)
    import csv

    rows = [[(cell or "").strip() for cell in row] for row in csv.reader(io.StringIO(text), dialect)]
    rows = [row for row in rows if any(row)]
    return f"=== CSV: {path.name} ===\n{_rows_to_text(rows)}"


def extract_docx(path: Path) -> str:
    document = Document(io.BytesIO(path.read_bytes()))
    sections: list[str] = []
    paragraphs = _paragraph_text(document)
    if paragraphs:
        sections.append("=== Paragraphs ===\n" + "\n".join(paragraphs))
    for index, table in enumerate(document.tables, start=1):
        text = _table_to_text(table)
        if text.strip():
            sections.append(f"=== Table {index} ===\n{text}")
    return "\n\n".join(sections)


EXTRACTORS: dict[str, tuple[str, Extractor]] = {
    "pdf": ("*.pdf", extract_pdf),
    "xlsx": ("*.xlsx", extract_xlsx),
    "csv": ("*.csv", extract_csv),
    "docx": ("*.docx", extract_docx),
}


def flatten(obj: Any, prefix: str = "") -> dict[str, Any]:
    out: dict[str, Any] = {}
    if isinstance(obj, dict):
        for key, value in obj.items():
            out.update(flatten(value, f"{prefix}.{key}" if prefix else key))
    elif isinstance(obj, list):
        for index, value in enumerate(obj):
            out.update(flatten(value, f"{prefix}[{index}]"))
    else:
        out[prefix] = obj
    return out


def values_match(left: Any, right: Any) -> bool:
    if left is None and right is None:
        return True
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        return math.isclose(float(left), float(right), rel_tol=1e-3, abs_tol=0.01)
    return str(left).strip() == str(right).strip()


def accuracy(predicted: dict[str, Any], expected: dict[str, Any]) -> tuple[float, list[str]]:
    predicted_flat = flatten(predicted)
    expected_flat = flatten(expected)
    keys = sorted(set(predicted_flat) | set(expected_flat))
    mismatches = [
        key for key in keys
        if not values_match(predicted_flat.get(key), expected_flat.get(key))
    ]
    score = (len(keys) - len(mismatches)) / len(keys) if keys else 1.0
    return score, mismatches


def fixture_cases(root: Path, formats: list[str], sample_per_format: int | None) -> list[tuple[str, Path]]:
    cases: list[tuple[str, Path]] = []
    for fmt in formats:
        pattern, _ = EXTRACTORS[fmt]
        files = sorted((root / fmt).glob(pattern))
        if sample_per_format:
            files = files[:sample_per_format]
        cases.extend((fmt, path) for path in files)
    return cases


def run_offline_case(fmt: str, path: Path, profile: dict[str, Any]) -> dict[str, Any]:
    expected_path = path.with_name(f"{path.stem}.expected.json")
    expected = json.loads(expected_path.read_text(encoding="utf-8"))
    schema_errors = list(Draft202012Validator(profile["schema"]).iter_errors(expected))
    pdf_classification = classify(path.read_bytes()).to_dict() if fmt == "pdf" else None
    text = ""
    text_check_skipped = bool(pdf_classification and pdf_classification["is_scanned"])
    if not text_check_skipped:
        text = EXTRACTORS[fmt][1](path)
    report = {
        "format": fmt,
        "fixture": str(path.relative_to(REPO_ROOT)),
        "expected_json_valid": not schema_errors,
        "text_length": len(text),
        "text_non_empty": bool(text.strip()),
        "text_check_skipped": text_check_skipped,
    }
    if pdf_classification:
        report["pdf_classification"] = pdf_classification
    return report


def run_llm_case(fmt: str, path: Path, profile: dict[str, Any], client: OpenAIExtractionClient) -> dict[str, Any]:
    text = EXTRACTORS[fmt][1](path)
    expected_path = path.with_name(f"{path.stem}.expected.json")
    expected = json.loads(expected_path.read_text(encoding="utf-8"))
    context = {
        "client_id": "eval",
        "document_id": path.stem,
        "metadata_json": to_metadata_json({"fixture": str(path)}),
        "document_text": wrap_untrusted_document_text(text),
    }
    result = client.extract(profile=profile, document_text=text, context=context)
    score, mismatches = accuracy(result["data"], expected)
    schema_errors = list(Draft202012Validator(profile["schema"]).iter_errors(result["data"]))
    return {
        "format": fmt,
        "fixture": str(path.relative_to(REPO_ROOT)),
        "accuracy": round(score, 4),
        "mismatch_count": len(mismatches),
        "mismatches": mismatches[:25],
        "schema_valid": not schema_errors,
        "response_id": result["response_id"],
        "model": result["model"],
        "usage": result["usage"],
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", choices=("offline", "llm"), default="offline")
    parser.add_argument("--profile", default="payroll")
    parser.add_argument("--version", default="v1")
    parser.add_argument("--fixtures-root", type=Path, default=REPO_ROOT / "tests/fixtures/payroll")
    parser.add_argument("--formats", default="pdf,xlsx,csv,docx")
    parser.add_argument("--sample-per-format", type=int, default=None)
    parser.add_argument("--min-accuracy", type=float, default=0.95)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()

    formats = [fmt.strip() for fmt in args.formats.split(",") if fmt.strip()]
    unknown = [fmt for fmt in formats if fmt not in EXTRACTORS]
    if unknown:
        raise SystemExit(f"Unknown formats: {unknown}")

    profile = load_profile(str(REPO_ROOT / "profiles"), args.profile, args.version)
    cases = fixture_cases(args.fixtures_root, formats, args.sample_per_format)
    client = OpenAIExtractionClient() if args.mode == "llm" else None

    results = [
        run_llm_case(fmt, path, profile, client) if client else run_offline_case(fmt, path, profile)
        for fmt, path in cases
    ]
    report = {
        "mode": args.mode,
        "profile": f"{args.profile}/{args.version}",
        "case_count": len(results),
        "results": results,
    }

    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(json.dumps(report, ensure_ascii=True, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=True, indent=2))

    if args.mode == "offline":
        failed = [
            r for r in results
            if not r["expected_json_valid"] or (not r["text_check_skipped"] and not r["text_non_empty"])
        ]
    else:
        failed = [
            r for r in results
            if not r["schema_valid"] or r["accuracy"] < args.min_accuracy
        ]
    if failed:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
